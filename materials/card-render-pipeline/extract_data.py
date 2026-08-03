"""
Convergence card-render pipeline — data extractor.

Reads the master stat sheet (xlsx) and writes data/cards.json: a clean,
render-ready record for every card, plus the rarity -> frame-colour mapping
the rulebook defines (Rare=Black, Epic=Purple, Legendary=Yellow, Mythic=Red).

Run:  py -3.14 extract_data.py
"""
import json, os, re
from pathlib import Path
import openpyxl

HERE = Path(__file__).resolve().parent
MATERIALS = HERE.parent
XLSX = MATERIALS / "Convergence card stat excel sheet.xlsx"

# Rulebook v1.0: "Rarity – Rare (Black), Epic (Purple), Legendary (Yellow), Mythic (Red)"
RARITY_COLOR = {
    "Rare": "black",
    "Epic": "purple",
    "Legendary": "yellow",
    "Mythic": "red",
}

def clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    # normalise the curly quotes the sheet mixes in
    s = s.replace("“", '"').replace("”", '"').replace("’", "'")
    return s

def as_int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Cards"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [clean(c) for c in rows[0]]
    cards = []
    for r in rows[1:]:
        if r[0] is None:
            continue
        d = {hdr[i]: r[i] for i in range(len(hdr))}
        cost = as_int(d.get("Cost"))
        name = clean(d.get("Name"))
        if not name or cost is None:
            continue
        rarity = clean(d.get("Rarity"))
        cards.append({
            "name": name,
            "cost": cost,
            "atk": as_int(d.get("ATK")),
            "hp": as_int(d.get("HP")),
            "rarity": rarity,
            "color": RARITY_COLOR.get(rarity, "black"),
            "camp": clean(d.get("Camp")),
            "alignment": clean(d.get("Alignment")),
            "effect": clean(d.get("Card Effect")),
            "flavor": clean(d.get("Flavor Text")).strip('"'),
            "origin": clean(d.get("Origin")),
        })
    # --- Ascension Relics (separate sheet: no ATK/HP/camp/alignment/rarity) ---
    wsr = wb["Relics"]
    rrows = list(wsr.iter_rows(values_only=True))
    rhdr = [clean(c) for c in rrows[0]]
    n_relics = 0
    for r in rrows[1:]:
        d = {rhdr[i]: r[i] for i in range(len(rhdr))}
        name = clean(d.get("Name"))
        if not name:
            continue
        eff = clean(d.get("Card Effect"))
        if eff == "-":
            eff = ""
        cards.append({
            "name": name,
            "cost": as_int(d.get("Cost")),
            "atk": None, "hp": None,
            "rarity": "Relic", "color": "relic",
            "camp": "ASCENSION", "alignment": "RELIC",
            "effect": eff,
            "flavor": clean(d.get("Flavor Text")).strip('"'),
            "origin": clean(d.get("Origin")),
            "type": "relic",
        })
        n_relics += 1

    out = {"rarity_color": RARITY_COLOR, "cards": cards}
    (HERE / "data").mkdir(exist_ok=True)
    with open(HERE / "data" / "cards.json", "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"Wrote {len(cards)} cards ({n_relics} relics) -> data/cards.json")
    # quick sanity: 10-mana breakdown
    tens = [c for c in cards if c["cost"] == 10]
    from collections import Counter
    print("10-mana colours:", dict(Counter(c["color"] for c in tens)))

if __name__ == "__main__":
    main()
